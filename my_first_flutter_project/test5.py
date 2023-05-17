import http.client
import json
import pandas as pd
import pickle as pk
from urllib.parse import urlencode
from datetime import datetime, timedelta
import time
from collections import defaultdict
import math
import pandas as pd
import heapq as hq
import heapq
import numpy as np
import openpyxl 
from openpyxl import Workbook, load_workbook
import numpy as np
import networkx as nx


df=pd.read_excel("C:\\Users\\Nikhil\\Desktop\\dpWorld\\data\\3monthdata.xlsx", sheet_name=None)
all_df = pd.concat(df, ignore_index=True)
pd.set_option('display.max_rows', None)
all_df.groupby(['pickup_branch_code'])['number_of_packages'].sum().sort_values(ascending=False)
all_df.groupby(['delivery_branch_code'],sort=True)['number_of_packages'].sum().sort_values(ascending=False)

import http.client
import json
import pandas as pd
import pickle as pk
from urllib.parse import urlencode
from datetime import datetime, timedelta
import time
from collections import defaultdict
import math
import pandas as pd
import heapq as hq
import heapq
import numpy as np
import openpyxl 
from openpyxl import Workbook, load_workbook

_url="/maps/api/distancematrix/json?"
_sec="S43nRiurSxoc9z39UYwkBXCBfzgW6"
conn = http.client.HTTPSConnection(host="api.distancematrix.ai")


df=pd.read_excel("C:\\Users\\Nikhil\\Desktop\\dpWorld\\data\\3monthdata.xlsx", sheet_name=None)
df1=pd.read_csv("C:\\Users\\Nikhil\\Desktop\\dpWorld\\data\\location_data.csv")
df2=pd.read_csv("C:\\Users\\Nikhil\\Desktop\\dpWorld\\data\\one_day_data.csv")
df3=pd.read_excel("C:\\Users\\Nikhil\\Desktop\\dpWorld\\data\\Top_20.xlsx")
df4=pd.read_csv("C:\\Users\\Nikhil\\Desktop\\dpWorld\\data\\Top_20_branches.csv")
df5=pd.read_csv("C:\\Users\\Nikhil\\Desktop\\dpWorld\\data\\Top_20_Location.csv")
df6=pd.read_csv("C:\\Users\\Nikhil\\Desktop\\dpWorld\\data\\Top_20_Packages_Weight.csv")
all_df = pd.concat(df, ignore_index=True)
pd.set_option('display.max_rows', None)
all_df['packages_weight']=all_df['number_of_packages']*all_df['actual_weight']
df6['lat'].apply(lambda x: float(x))
df6['long'].apply(lambda x: float(x))



wb=Workbook()
ws=wb.active
ws['A1'].value='RouteID'
ws['B1'].value='RouteName'
ws['C1'].value='LegOrg(x)'
ws['D1'].value='LegOrg(y)'
ws['E1'].value='Departure'
ws['F1'].value='Arrival'
ws['G1'].value='Running Halt'
ws['H1'].value='Halt Time'
ws['I1'].value='Distance'
ws['J1'].value='Total Distance'
ws['K1'].value='Total Run Hrs'
ws['L1'].value='Route Type'
ws['M1'].value='PayLoad'
ws['N1'].value='Vehicle Size'
ws['O1'].value='Vehicle Count'

def distance(origin, destination):
    lat1, lon1=origin
    lat2, lon2=destination
    radius=6371
    dlat=math.radians(lat2-lat1)
    dlon=math.radians(lon2-lon1)
    a=math.sin(dlat/2)*math.sin(dlat/2)+math.cos(math.radians(lat1)) \
        *math.cos(math.radians(lat2))*math.sin(dlon/2)*math.sin(dlon/2)
    c=2*math.atan2(math.sqrt(a),math.sqrt(1-a))
    d=radius*c
    return d

def helper(x):
    return x

my_set=set([])
for i in df6.index:
    my_set.add(df6['Branches'][i]) 
    
print(my_set)
print(len(my_set))
my_set1=set([])
list1=[]
for val in my_set:
    list1.append(val)
print(list1)
print(len(list1))
Dict1={}
for i in my_set:
    origin_lat=df6[df6['Branches']==i].lat
    origin_long=df6[df6['Branches']==i].long
    origin=(origin_lat,origin_long)
    list2=[]
    for j in df6.index:
        b=df6['Branches'][j]
        dest_lat=df6['lat'][j]
        dest_long=df6['long'][j]
        destination=(dest_lat,dest_long)
        d=df6['number_weight_packages'][j]
        list11=[]
        if ((distance(origin,destination)>150) and (distance(origin,destination)<750)):
            list2.append(b)
             
                                
    Dict={}    
    if len(list2)>0:
        Dict[i]=list2
    if i=='LUHB':
        Dict[i]+=['IXCB']
    if i=='IXCB':
        Dict[i]+=['LUHB']
    if i=='NAGB':
        Dict[i]+=['CCUB']
    if i=='CCUB':
        Dict[i]+=['NAGB']
    print(Dict)
    Dict1.update(Dict)    


Dict5={}
count=0
for i in df6.index:
    Dict5[df6['Branch_Number'][i]]=df6['Branches'][i]
print(Dict5)
Dict6={}
Dict7={}
Dict8={}
count=0
for i in df6.index:
    Dict6[df6['Branches'][i]]=df6['Branch_Number'][i]
    Dict7[df6['Branches'][i]]=df6['number_weight_packages'][i]
    Dict8[df6['number_weight_packages'][i]]=df6['Branches'][i]
    
print(Dict6)
print(len(Dict6))
adj_list =[[] for _ in range(22)]
for i in Dict1:
    list8=[]
    list8=Dict1.get(i)
    for j in list8:
        adj_list[Dict6.get(i)].append(j)
        
    
print(adj_list)
print(len(adj_list))

count0=1
for _key in df6.index:
    for _x in df6.index:
        src= df6['Branches'][_key]
        dest= df6['Branches'][_x]
        x=df6[df6['Branches']==src]['number_weight_packages']
        print(x)
        max_heap=[(Dict6.get(src),src,dest,x)]
        print(max_heap)
        seen=set()
        a=False
        list12=[]
        origin_lat2=(float)(df6[df6['Branches']==src].lat)
        origin_long2=(float)(df6[df6['Branches']==src].long)
        origin2=(origin_lat,origin_long)
        dest_lat=(float)(df6[df6['Branches']==dest].lat)
        dest_long=(float)(df6[df6['Branches']==dest].long)
        destination=(dest_lat,dest_long)
        list13=[]
        if  distance(origin2,destination)>=1500:
            continue
        while max_heap:
            curr=hq.heappop(max_heap)
            if curr[1] in seen:
                continue
            list12.append(curr[1])
            if curr[1]==curr[2]:
                a=True 
                break
            seen.add(curr[1])
            list13.append(curr[1])
            origin_lat=(float)(df6[df6['Branches']==curr[1]].lat)
            origin_long=(float)(df6[df6['Branches']==curr[1]].long)
            origin=(origin_lat,origin_long)
            list3=[]
            for u in adj_list[curr[0]]:
                list3.append(u)
            for u in adj_list[curr[0]]:
                if u not in seen:
                    origin_lat1=(float)(df6[df6['Branches']==u].lat)
                    origin_long1=(float)(df6[df6['Branches']==u].long)
                    origin1=(origin_lat1,origin_long1)
                    if (distance(origin,destination)<distance(origin1,destination)) :
                        continue
                    else:
                        hq.heappush(max_heap,(Dict6.get(u),u,curr[2],Dict7.get(u)))
                        list13.append(u)
        print(a)
        print(Dict7)
        print(list12)
        print(list13)

        for i in list12:
            for j in list12:
                if i==j:
                    continue
                else:
                    origin_lat=(float)(df6[df6['Branches']==i].lat)
                    origin_long=(float)(df6[df6['Branches']==i].long)
                    dest_lat=(float)(df6[df6['Branches']==j].lat)
                    dest_long=(float)(df6[df6['Branches']==j].long)
                    origin=(origin_lat,origin_long)
                    destination=(dest_lat,dest_long)
                    if distance(origin,destination)<75:
                        d=min(Dict7.get(i),Dict7.get(j))
                        list12.remove(Dict8.get(d))

                        
                        
        print(list12)
        n=len(list12)
        print(n)
        stack=[]
        list4=[]
        for i in list12:
            if i==src or i==dest:
                continue
            list4.append(i)
        print(list4)
        print(len(list4))
        n=len(list4)
        for i in range(0,n):
            stack.append(list4[i])
            if i==0:
                continue
            j=i
            while len(stack)>1 and j>0  and  stack[-1] not in Dict1.get(stack[-2]):
                stack.pop()
                
        
        print(stack[0:len(stack)])
        stack1=[]
        while stack:
            stack1.append(stack.pop())
        print(stack)
        print(stack1)
        list12=[]
        list12.append(src)
        while stack1:
            list12.append(stack1.pop())
        list12.append(dest)
        print(list12)
        print(len(list12))  
        my_set=set()
        print(my_set)
        n=len(list12)
        stack=[]
        n=len(list12)
        for i in range(0,n):
            stack.append(list12[i])
            if i==0:
                continue
            while len(stack)>1 :
                origin_lat1=(float)(df6[df6['Branches']==stack[-2]].lat)
                origin_long1=(float)(df6[df6['Branches']==stack[-2]].long)
                origin1=(origin_lat1,origin_long1)
                origin_lat2=(float)(df6[df6['Branches']==stack[-1]].lat)
                origin_long2=(float)(df6[df6['Branches']==stack[-1]].long)
                origin2=(origin_lat2,origin_long2)
                dest_lat=(float)(df6[df6['Branches']==dest].lat)
                dest_long=(float)(df6[df6['Branches']==dest].long)
                destination=(dest_lat,dest_long)
                if distance(origin1,destination)>distance(origin2,destination):
                    break
                else:
                    stack.pop()
        print(stack[0:len(stack)])
        stack1=[]
        while stack:
            stack1.append(stack.pop())
        list12=[]

        while stack1:
            list12.append(stack1.pop())

        print(list12)
        print(len(list12))

        for i in list12:
            if i==dest:
                for j in list12:
                    if j==src or j==dest:
                        continue
                    origin_lat=(float)(df6[df6['Branches']==i].lat)
                    origin_long=(float)(df6[df6['Branches']==i].long)
                    dest_lat=(float)(df6[df6['Branches']==j].lat)
                    dest_long=(float)(df6[df6['Branches']==j].long)
                    origin=(origin_lat,origin_long)
                    destination=(dest_lat,dest_long)
                    if distance(origin,destination)<75:
                        list12.remove(j)
                                
        stack=[]
        
        for i in list12:
            stack.append(i)
            if i==dest:
                continue
            if len(stack)>2:
                origin_lat=(float)(df6[df6['Branches']==stack[-2]].lat)
                origin_long=(float)(df6[df6['Branches']==stack[-2]].long)
                origin_lat1=(float)(df6[df6['Branches']==stack[-1]].lat)
                origin_long1=(float)(df6[df6['Branches']==stack[-1]].long)
                dest_lat=(float)(df6[df6['Branches']==dest].lat)
                dest_long=(float)(df6[df6['Branches']==dest].long)
                origin=(origin_lat,origin_long)
                origin1=(origin_lat1,origin_long1)
                destination=(dest_lat,dest_long)
                if (distance(origin,destination)-distance(origin1,destination)<250):
                    stack.pop()
                else:
                    continue
                
                
        list12=[]
        stack1=[]
        while stack:
            stack1.append(stack.pop())
        while stack1:
            list12.append(stack1.pop())
        print(list12)
        print(len(list12))
        
        res_list=[[]]
        list2=[]
        for i in list12:
            list2.append(i)
            if i==dest or len(list2)==4:
                res_list.append(list2)
                list2=[]
                list2.append(i)
        res_list.remove([])
        print(res_list)
        print(len(res_list))

        b_dist=False
        a_dist=False
        n=len(list12)
        for i in range(0,len(list12)-1):
            origin_lat2=(float)(df6[df6['Branches']==list12[i]].lat)
            origin_long2=(float)(df6[df6['Branches']==list12[i]].long)
            origin2=(origin_lat2,origin_long2)
            dest_lat=(float)(df6[df6['Branches']==list12[i+1]].lat)
            dest_long=(float)(df6[df6['Branches']==list12[i+1]].long)
            destination=(dest_lat,dest_long)
            if distance(origin2,destination)>800:
                b_dist=True
                
        
        for i in range(0,len(list12)-2):
            if i>=n-2:
                continue
            origin_lat_u=(float)(df6[df6['Branches']==list12[i]].lat)
            origin_long_u=(float)(df6[df6['Branches']==list12[i]].long)
            origin_u=(origin_lat_u,origin_long_u)
            origin_lat_v=(float)(df6[df6['Branches']==list12[i+1]].lat)
            origin_long_v=(float)(df6[df6['Branches']==list12[i+1]].long)
            origin_v=(origin_lat_v,origin_long_v)
            dest_lat_u=(float)(df6[df6['Branches']==list12[i+2]].lat)
            dest_long_u=(float)(df6[df6['Branches']==list12[i+2]].long)
            destination_u=(dest_lat_u,dest_long_u)
            if distance(origin_u,origin_v)>distance(origin_u,destination_u):
                a_dist=True
            if a_dist:
                temp=list12[i+1]
                list12[i+1]=list12[i+2]
                list12[i+2]=temp
        if b_dist==True:
            continue

        _sec="S43nRiurSxoc9z39UYwkBXCBfzgW6"
        def dist_map_api(coord1, coord2):
            _org = f"{coord1[0]},{coord1[1]}"
            _dest = f"{coord2[0]},{coord2[1]}" 
            try:
                conn.request(method='GET', url=f"{_url}{urlencode({'origins': _org, 'destinations': _dest, 'key': _sec})}")
                res01 = conn.getresponse()
                _data = json.loads(res01.read().decode('utf-8'))
                return [_data["rows"][0]["elements"][0]["distance"]["value"], _data["rows"][0]["elements"][0]["duration"]["value"]]
            except Exception as e:
                return [0,0]
           
       
            
        print(list12)
        print(dist_map_api((30.70460455,76.80458955),(28.8361342,76.8457048)))
        count=count0
        n=len(list12)
        count1=ws.max_row
        lat0=(float)(df6[df6['Branches']==src].lat)
        long0=(float)(df6[df6['Branches']==src].long)


        count1=ws.max_row

        origin_lat=(float)(df6[df6['Branches']==src].lat)
        origin_long=(float)(df6[df6['Branches']==src].long)
        origin=(origin_lat,origin_long)
        dest_lat=(float)(df6[df6['Branches']==dest].lat)
        dest_long=(float)(df6[df6['Branches']==dest].long)
        destination=(dest_lat,dest_long)
        n=len(list12)
        dest=list12[n-1]
        if len(list12)>=3 and len(list12)<=5 and distance(origin,destination)<=1100 :
            dist_sum=0
            for i in range(0,n-1):
                lat1=(float)(df6[df6['Branches']==list12[i]].lat)
                long1=(float)(df6[df6['Branches']==list12[i]].long)
                lat2=(float)(df6[df6['Branches']==list12[i+1]].lat)
                long2=(float)(df6[df6['Branches']==list12[i+1]].long)
                list2=[]
                list2=dist_map_api((lat1,long1),(lat2,long2))
                dist_sum=dist_sum+list2[0]
            lat1=(float)(df6[df6['Branches']==list12[0]].lat)
            long1=(float)(df6[df6['Branches']==list12[0]].long)
            lat2=(float)(df6[df6['Branches']==list12[n-1]].lat)
            long2=(float)(df6[df6['Branches']==list12[n-1]].long)
            list2=[]
            list2=dist_map_api((lat1,long1),(lat2,long2))
            src_dest_dist=(int)(list2[0]/1000)
            src_dest_time=(int)(list2[1]/3600)
            dist_sum=(int)(dist_sum/1000)  
            for i in range(0,n-1):
                if dist_sum>1.8*src_dest_dist:
                    if distance(origin,destination)>800:
                        break
                    count1=ws.max_row
                    count1=count1+1
                    count0=count0+1
                    ws.row_dimensions[count1].height=50
                    ws['C'+str(count1)].value=list12[i]
                    ws['D'+str(count1)].value=list12[n-1]
                    ws['B'+str(count1)].value=list12[0]+"-"+list12[n-1]
                    ws['G'+str(count1)].value=src_dest_time
                    ws['K'+str(count1)].value=src_dest_time
                    ws['I'+str(count1)].value=src_dest_dist
                    ws['J'+str(count1)].value=src_dest_dist
                    ws['A'+str(count1)].value=count
                    break
                    
                if i==0:
                    count0=count0+1
                    count1=count1+1
                    ws.row_dimensions[count1].height=70
                    a=list12[i]
                    b=list12[n-1]
                    lat1=(float)(df6[df6['Branches']==a].lat)
                    long1=(float)(df6[df6['Branches']==a].long)
                    lat2=(float)(df6[df6['Branches']==b].lat)
                    long2=(float)(df6[df6['Branches']==b].long)
                    list2=[]
                    list2=dist_map_api((lat1,long1),(lat2,long2))
                    ws['C'+str(count1)].value=list12[0]
                    ws['D'+str(count1)].value=list12[n-1]
                    x=""
                    for j in range(1,n):
                        x=a+"-"+list12[j]
                        a=x
                    ws['B'+str(count1)].value=x
                    ws['G'+str(count1)].value=src_dest_time
                    ws['K'+str(count1)].value=src_dest_time
                    ws['I'+str(count1)].value=src_dest_dist
                    ws['J'+str(count1)].value=src_dest_dist
                    ws['A'+str(count1)].value=count
                    if len(list12)==2:
                        break
                lat1=(float)(df6[df6['Branches']==list12[i]].lat)
                long1=(float)(df6[df6['Branches']==list12[i]].long)
                lat2=(float)(df6[df6['Branches']==list12[i+1]].lat)
                long2=(float)(df6[df6['Branches']==list12[i+1]].long)
                list2=[]
                list2=dist_map_api((lat1,long1),(lat2,long2))
                count1=count1+1
                ws.row_dimensions[count1].height=50
                ws['C'+str(count1)].value=list12[i]
                ws['D'+str(count1)].value=list12[i+1]
                ws['B'+str(count1)].value=list12[i]+"-"+list12[i+1]
                ws['G'+str(count1)].value=(int)(list2[1]/3600)
                ws['K'+str(count1)].value=(int)(list2[1]/3600)
                ws['I'+str(count1)].value=(int)(list2[0]/1000)
                ws['J'+str(count1)].value=(int)(list2[0]/1000)
                            
            wb.save("C:\\Users\\Nikhil\\Desktop\\dpWorld\\data\\Routes10.xlsx")
            
            
            

