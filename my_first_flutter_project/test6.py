import http.client
import json
import pandas as pd
import pickle as pk
from urllib.parse import urlencode
from datetime import datetime, timedelta
import time
from collections import defaultdict
import math
import pandas as pd
import heapq as hq
import heapq
import numpy as np
import openpyxl         
from openpyxl import Workbook, load_workbook
import sys
_url="/maps/api/distancematrix/json?"
_sec="rDP0yykiOrLfwM6ranORSmkPSkwhj"
conn = http.client.HTTPSConnection(host="api.distancematrix.ai")


df=pd.read_excel("C:\\Users\\Nikhil\\Desktop\\dpWorld\\data\\3monthdata.xlsx", sheet_name=None)
df1=pd.read_csv("C:\\Users\\Nikhil\\Desktop\\dpWorld\\data\\location_data.csv")
df2=pd.read_csv("C:\\Users\\Nikhil\\Desktop\\dpWorld\\data\\one_day_data.csv")
df3=pd.read_excel("C:\\Users\\Nikhil\\Desktop\\dpWorld\\data\\Top_20.xlsx")
df4=pd.read_csv("C:\\Users\\Nikhil\\Desktop\\dpWorld\\data\\Top_20_branches.csv")
df5=pd.read_csv("C:\\Users\\Nikhil\\Desktop\\dpWorld\\data\\Top_20_Location.csv")
df6=pd.read_csv("C:\\Users\\Nikhil\\Desktop\\dpWorld\\data\\Top_30_Packages_Weight.csv")
all_df = pd.concat(df, ignore_index=True)
pd.set_option('display.max_rows', None)
df6['lat'].apply(lambda x: float(x))
df6['long'].apply(lambda x: float(x))

wb=Workbook()
ws=wb.active
ws['A1'].value='RouteID'
ws['B1'].value='RouteName'
ws['C1'].value='LegOrg(x)'
ws['D1'].value='LegOrg(y)'
ws['E1'].value='Departure'
ws['F1'].value='Arrival'
ws['G1'].value='Running Halt'
ws['H1'].value='Halt Time'
ws['I1'].value='Distance'
ws['J1'].value='Total Distance'
ws['K1'].value='Total Run Hrs'
ws['L1'].value='Route Type'
ws['M1'].value='PayLoad'
ws['N1'].value='Vehicle Size'
ws['O1'].value='Vehicle Count'



Dict3={}
def distance(origin, destination):
    lat1, lon1=origin
    lat2, lon2=destination
    radius=6371
    dlat=math.radians(lat2-lat1)
    dlon=math.radians(lon2-lon1)
    a=math.sin(dlat/2)*math.sin(dlat/2)+math.cos(math.radians(lat1)) \
        *math.cos(math.radians(lat2))*math.sin(dlon/2)*math.sin(dlon/2)
    c=2*math.atan2(math.sqrt(a),math.sqrt(1-a))
    d=radius*c
    return d

_sec="rDP0yykiOrLfwM6ranORSmkPSkwhj"
def dist_map_api(coord1, coord2):
    _org = f"{coord1[0]},{coord1[1]}"
    _dest = f"{coord2[0]},{coord2[1]}" 
    try:
        conn.request(method='GET', url=f"{_url}{urlencode({'origins': _org, 'destinations': _dest, 'key': _sec})}")
        res01 = conn.getresponse()
        _data = json.loads(res01.read().decode('utf-8'))
        return [_data["rows"][0]["elements"][0]["distance"]["value"], _data["rows"][0]["elements"][0]["duration"]["value"]]
    except Exception as e:
        return [0,0]
           

Dict1={}
for i in df6.index:
    a=df6["Branches"][i]
    origin_lat=df6["lat"][i]
    origin_long=df6["long"][i]
    origin=(origin_lat,origin_long)
    list2=[]
    for j in df6.index:
        b=df6['Branches'][j]
        dest_lat=df6['lat'][j]
        dest_long=df6['long'][j]
        destination=(dest_lat,dest_long)
        d=df6['number_weight_packages'][j]
        list11=[]
        if ((distance(origin,destination)>70) and (distance(origin,destination)<565)):
            list2.append(b)
    Dict={}
    if len(list2)>0:
        Dict[a]=list2
    Dict1.update(Dict)
    print(Dict)
    

Dict5={}
Dict6={}
Dict7={}
Dict8={}
count=0
for i in df6.index:
    Dict5[df6['Branch_Number'][i]]=df6['Branches'][i]
    Dict6[df6['Branches'][i]]=df6['Branch_Number'][i]
    Dict7[df6['Branches'][i]]=df6['number_weight_packages'][i]
    Dict8[df6['number_weight_packages'][i]]=df6['Branches'][i]
    
adj_list =[[] for _ in range(42)]
for i in Dict1:
    list8=[]
    list8=Dict1.get(i)
    for j in list8:
        adj_list[Dict6.get(i)].append(j)

print(adj_list)
print(len(adj_list))
print(Dict1)
count0=1
count7=1
Routes={}
R_count=0
s=0
for _key in df6.index:
    for _x in df6.index:
        count7=count7+1
        src= df6['Branches'][_key]
        dest= df6['Branches'][_x]
        x=df6[df6['Branches']==src]['number_weight_packages']
        print(x)
        max_heap=[(Dict6.get(src),src,dest,x)]
        print(max_heap)
        seen=set()
        a=False
        list12=[]
        list13=[]
        origin_lat2=(float)(df6[df6['Branches']==src].lat)
        origin_long2=(float)(df6[df6['Branches']==src].long)
        origin2=(origin_lat2,origin_long2)
        dest_lat=(float)(df6[df6['Branches']==dest].lat)
        dest_long=(float)(df6[df6['Branches']==dest].long)
        destination=(dest_lat,dest_long)
        if distance(origin2,destination)>=965 or distance(origin2,destination)<=175:
            continue
        a=False
        print(a)
        while max_heap:
            curr=hq.heappop(max_heap)
            if curr[1] in seen:
                continue
            list12.append(curr[1])
            if curr[1]==curr[2]:
                a=True 
                break
            seen.add(curr[1])
            list13.append(curr[1])
            origin_lat=(float)(df6[df6['Branches']==curr[1]].lat)
            origin_long=(float)(df6[df6['Branches']==curr[1]].long)
            origin=(origin_lat,origin_long)
            list3=[]
            for u in adj_list[curr[0]]:
                if u not in seen:
                    origin_lat1=(float)(df6[df6['Branches']==u].lat)
                    origin_long1=(float)(df6[df6['Branches']==u].long)
                    origin1=(origin_lat1,origin_long1)
                    if (distance(origin,destination)<distance(origin1,destination)) :
                        continue
                    list3.append(u)
                
                
            for u in adj_list[curr[0]]:
                if u not in seen:
                    origin_lat1=(float)(df6[df6['Branches']==u].lat)
                    origin_long1=(float)(df6[df6['Branches']==u].long)
                    origin1=(origin_lat1,origin_long1) 
                    if (distance(origin,destination)<distance(origin1,destination)) or distance(origin1,destination)+distance(origin,origin1)>1.1*(distance(origin,destination)):
                        continue
                    else:
                        hq.heappush(max_heap,(Dict6.get(u),u,curr[2],Dict7.get(u)))
                        list13.append(u)
        print(list12)
        print("list12_#1")
        print(a)
        for i in list12:
            for j in list12:
                if i==j:
                    continue
                else:
                    origin_lat=(float)(df6[df6['Branches']==i].lat)
                    origin_long=(float)(df6[df6['Branches']==i].long)
                    dest_lat=(float)(df6[df6['Branches']==j].lat)
                    dest_long=(float)(df6[df6['Branches']==j].long)
                    origin=(origin_lat,origin_long)
                    destination=(dest_lat,dest_long)
                    if distance(origin,destination)<70:
                        d=min(Dict7.get(i),Dict7.get(j))
                        if Dict8.get(d) not in list12:
                            continue
                        list12.remove(Dict8.get(d))
        print(list12)
        print("list12_#2")
        n=len(list12)
        print(n)
        stack=[]
        list4=[]
        for i in list12:
            if i==src or i==dest:
                continue
            list4.append(i)
        print(list4)
        print(len(list4))
        n=len(list4)
        for i in range(0,n):
            stack.append(list4[i])
            if i==0:
                continue
            j=i
            while len(stack)>1 and j>0  and  stack[-1] not in Dict1.get(stack[-2]):
                stack.pop()
                
        print(stack[0:len(stack)])
        stack1=[]
        while stack:
            stack1.append(stack.pop())
        print(stack)
        print(stack1)
        list12=[]
        list12.append(src)
        while stack1:
            list12.append(stack1.pop())
        list12.append(dest)
        print(list12)
        print("list12_#3")
        print(len(list12))
        
        
        for i in list12:
            if i==dest:
                for j in list12:
                    if j==src or j==dest:
                        continue
                    origin_lat=(float)(df6[df6['Branches']==i].lat)
                    origin_long=(float)(df6[df6['Branches']==i].long)
                    dest_lat=(float)(df6[df6['Branches']==j].lat)
                    dest_long=(float)(df6[df6['Branches']==j].long)
                    origin=(origin_lat,origin_long)
                    destination=(dest_lat,dest_long)
                    if distance(origin,destination)<70:
                        list12.remove(j)
        print(list12)
        print("list12_#4")
        
        list4=[]
        for i in list12:
            list4.append(i)
        n=len(list4)
        stack=[]
        for i in range(0,n):
            stack.append(list4[i])
            if i==0 or i==n-1 or len(stack)<=1:
                continue
            j=i
            origin_lat1=(float)(df6[df6['Branches']==stack[-1]].lat)
            origin_long1=(float)(df6[df6['Branches']==stack[-1]].long)
            origin1=(origin_lat1,origin_long1)
            origin_lat2=(float)(df6[df6['Branches']==stack[-2]].lat)
            origin_long2=(float)(df6[df6['Branches']==stack[-2]].long)
            origin2=(origin_lat2,origin_long2)
            dest_lat=(float)(df6[df6['Branches']==dest].lat)
            dest_long=(float)(df6[df6['Branches']==dest].long)
            destination=(dest_lat,dest_long)
            a=distance(origin1,destination)
            b=distance(origin2,destination)
            while len(stack)>1 and a>=b:
                stack.pop()
                if len(stack)>1:
                    origin_lat1=(float)(df6[df6['Branches']==stack[-1]].lat)
                    origin_long1=(float)(df6[df6['Branches']==stack[-1]].long)
                    origin1=(origin_lat1,origin_long1)
                    origin_lat2=(float)(df6[df6['Branches']==stack[-2]].lat)
                    origin_long2=(float)(df6[df6['Branches']==stack[-2]].long)
                    origin2=(origin_lat2,origin_long2)
                    dest_lat=(float)(df6[df6['Branches']==dest].lat)
                    dest_long=(float)(df6[df6['Branches']==dest].long)
                    destination=(dest_lat,dest_long)
                    a=distance(origin1,destination)
                    b=distance(origin2,destination)
        print(stack[0:len(stack)])
        stack1=[]
        while stack:
            stack1.append(stack.pop())
        print(stack)
        print(stack1)
        list12=[]
        while stack1:
            list12.append(stack1.pop())
        
        print(list12)
        print("list12_#5")
        n=len(list12)
        
        dict_ins_loc={}
        for i in range(0,n-1):
            print(list12[i])
            list2=[]
            curr_loc=list12[i]
            origin_lat_u=(float)(df6[df6['Branches']==list12[i]].lat)
            origin_long_u=(float)(df6[df6['Branches']==list12[i]].long)
            origin_u=(origin_lat_u,origin_long_u)
            origin_lat_v=(float)(df6[df6['Branches']==list12[i+1]].lat)
            origin_long_v=(float)(df6[df6['Branches']==list12[i+1]].long)
            origin_v=(origin_lat_v,origin_long_v)
            for j in df6.index:
                b=df6['Branches'][j]
                dest_lat1=df6['lat'][j]
                dest_long1=df6['long'][j]
                destination1=(dest_lat,dest_long)
                if ((distance(origin_u,destination1)>70) and (distance(origin_u,destination1)<565)):
                    list2.append(b)
            
        
            print(list2)
            print("list2_")
            for u in list2:
                origin_lat_1=(float)(df6[df6['Branches']==u].lat)
                origin_long_1=(float)(df6[df6['Branches']==u].long)
                origin_1=(origin_lat_1,origin_long_1)   
                if distance(origin_u, origin_v)<distance(origin_1,origin_v) or distance(origin_u,origin1)>distance(origin_u,origin_v):
                    list2.remove(u)
            print(list2)
            print("list2")
            for u in list2:
                origin_lat_1=(float)(df6[df6['Branches']==u].lat)
                origin_long_1=(float)(df6[df6['Branches']==u].long)
                origin_1=(origin_lat_1,origin_long_1)
                if distance(origin_u,origin_1)+distance(origin_1,origin_v)<=1.05*distance(origin_u,origin_v) and (u not in list12):
                    if u not in Dict1.get(list12[i+1]):
                        continue
                    dict_ins_loc[u]=i+1
                    continue
            print(dict_ins_loc)
            print("dict_ins_loc")
                
            for u in dict_ins_loc:
                if u in list12:
                    continue
                else:
                    origin_lat_1=(float)(df6[df6['Branches']==u].lat)
                    origin_long_1=(float)(df6[df6['Branches']==u].long)
                    origin_1=(origin_lat_1,origin_long_1)
                    list12.insert(dict_ins_loc.get(u),u)
                    print(dict_ins_loc.get(u))
                    print(u)
                    print("u")
        
        
                    
        for i in list12:
            for j in list12:
                if i==j:
                    continue
                else:
                    origin_lat=(float)(df6[df6['Branches']==i].lat)
                    origin_long=(float)(df6[df6['Branches']==i].long)
                    dest_lat=(float)(df6[df6['Branches']==j].lat)
                    dest_long=(float)(df6[df6['Branches']==j].long)
                    origin=(origin_lat,origin_long)
                    destination=(dest_lat,dest_long)
                    if distance(origin,destination)<70:
                        d=min(Dict7.get(i),Dict7.get(j))
                        if Dict8.get(d) not in list12:
                            continue
                        list12.remove(Dict8.get(d))
        for i in list12:
            if i==dest:
                for j in list12:
                    if j==src or j==dest:
                        continue
                    origin_lat=(float)(df6[df6['Branches']==i].lat)
                    origin_long=(float)(df6[df6['Branches']==i].long)
                    dest_lat=(float)(df6[df6['Branches']==j].lat)
                    dest_long=(float)(df6[df6['Branches']==j].long)
                    origin=(origin_lat,origin_long)
                    destination=(dest_lat,dest_long)
                    if distance(origin,destination)<70:
                        list12.remove(j)
        print(list12)
        print("list12_#6")
        very_long=False
        n=len(list12)
        for i in range(0,n-1):
            origin_lat_u=(float)(df6[df6['Branches']==list12[i]].lat)
            origin_long_u=(float)(df6[df6['Branches']==list12[i]].long)
            origin_u=(origin_lat_u,origin_long_u)
            origin_lat_v=(float)(df6[df6['Branches']==list12[i+1]].lat)
            origin_long_v=(float)(df6[df6['Branches']==list12[i+1]].long)
            origin_v=(origin_lat_v,origin_long_v)
            if distance(origin_u,origin_v)>675:
                very_long=True
                    
    
    
        n=len(list12)
        for i in range(0,len(list12)-2):
            origin_lat_u=(float)(df6[df6['Branches']==list12[i]].lat)
            origin_long_u=(float)(df6[df6['Branches']==list12[i]].long)
            origin_u=(origin_lat_u,origin_long_u)
            origin_lat_v=(float)(df6[df6['Branches']==list12[i+1]].lat)
            origin_long_v=(float)(df6[df6['Branches']==list12[i+1]].long)
            origin_v=(origin_lat_v,origin_long_v)
            if (i+2)>=len(list12):
                continue
            origin_lat_1=(float)(df6[df6['Branches']==list12[i+2]].lat)
            origin_long_1=(float)(df6[df6['Branches']==list12[i+2]].long)
            origin_1=(origin_lat_1,origin_long_1)
            if distance(origin_u,origin_v)>distance(origin_u,origin_1) and list12[i+1] in Dict1.get(list12[i]) and list12[i+2] in Dict1.get(list12[i]):
                if list12[i+1] not in list12:
                    continue
                else:
                    list12.remove(list12[i+1])
            
        
             
        print(list12)
        print("list12 ending")
        if len(list12)<=2 :
            continue
        
        count1=1
        n=len(list12)
        for i in range(0,n-1):
            if list12[i] not in Dict3:
                continue
            if list12[i+1] in Dict3.get(list12[i]):
                count1=count1+1
        if n==count1:
            continue
        R_count=R_count+1
        Routes['r'+str(R_count)]=list12
        n=len(list12)
        m=len(list12)
        for val in range(0,m-1):
            _val_i=list12[val]
            _val_i1=list12[val+1]
            if _val_i not in Dict3:
                Dict3[_val_i]=[_val_i1]
            else:
                found=False
                for u in Dict3[_val_i]:
                    if u==_val_i1:
                        found=True
                if found==False:
                    Dict3[_val_i].append(_val_i1)
        s=s+1


print(Dict3)
for u in Dict3:
    list2=[]
    list2=Dict3.get(u)
    for v in list2:
        if v not in Dict3:
            continue
        if u not in Dict3.get(v):
            R_count=R_count+1
            Routes['r'+str(R_count)]=[v,u]
            if v not in Dict3:
                Dict3[v]=[u]
            else:
                Dict3[v].append(u)
                
                
for i in df6.index:
    l_loc=df6['Branches'][i]
    origin_lat_u=(float)(df6[df6['Branches']==l_loc].lat)
    origin_long_u=(float)(df6[df6['Branches']==l_loc].long)
    origin_u=(origin_lat_u,origin_long_u)
    for j in df6.index:
        s_loc=df6['Branches'][j]
        if l_loc==s_loc:
            continue
        origin_lat_v=(float)(df6[df6['Branches']==s_loc].lat)
        origin_long_v=(float)(df6[df6['Branches']==s_loc].long)
        origin_v=(origin_lat_v,origin_long_v)
        if l_loc not in Dict3:
            continue
        if distance(origin_u,origin_v)<=70 and s_loc not in Dict3.get(l_loc):
            R_count=R_count+1
            Routes['r'+str(R_count)]=[l_loc,s_loc]
            if l_loc not in Dict3:
                Dict3[l_loc]=[s_loc]
            else:
                Dict3[l_loc].append(s_loc)
         
df1=pd.read_csv("C:\\Users\\Nikhil\\Desktop\\dpWorld\\data\\location_data.csv")   
all_loc_list=[]
for i in df6.index:
    all_loc_list.append(df6['Branches'][i])
print(all_loc_list)
for i in df1.index:
    l_loc=df1['code'][i]
    origin_lat_u=(float)(df1[df1['code']==l_loc].lat)
    origin_long_u=(float)(df1[df1['code']==l_loc].long)
    origin_u=(origin_lat_u,origin_long_u)
    min_loc=""
    min=999999999999999999999
    if l_loc in all_loc_list:
        continue
    for j in df6.index:
        s_loc=df6['Branches'][j]
        if l_loc==s_loc:
            continue
        origin_lat_v=(float)(df6[df6['Branches']==s_loc].lat)
        origin_long_v=(float)(df6[df6['Branches']==s_loc].long)
        origin_v=(origin_lat_v,origin_long_v)
        if distance(origin_u,origin_v)<min:
            min=distance(origin_u,origin_v)
            min_loc=s_loc
    if min_loc!="":
        R_count=R_count+1
        Routes['r'+str(R_count)]=[l_loc,min_loc]
        R_count=R_count+1
        Routes['r'+str(R_count)]=[min_loc,l_loc]
        
        
for i in df6.index:
    a_loc=df6['Branches'][i]
    for j in df6.index:
        b_loc=df6['Branches'][j]
        if a_loc==b_loc:
            continue
        origin_lat_u=(float)(df1[df1['code']==a_loc].lat)
        origin_long_u=(float)(df1[df1['code']==a_loc].long)
        origin_u=(origin_lat_u,origin_long_u)
        origin_lat_v=(float)(df6[df6['Branches']==b_loc].lat)
        origin_long_v=(float)(df6[df6['Branches']==b_loc].long)
        origin_v=(origin_lat_v,origin_long_v)
        if b_loc not in Dict3:
            continue
        if distance(origin_u,origin_v)<=200 and a_loc not in Dict3.get(b_loc):
            conne=False
            for k in df6.index:
                c_loc=df6['Branches'][k]
                origin_lat_1=(float)(df6[df6['Branches']==c_loc].lat)
                origin_long_1=(float)(df6[df6['Branches']==c_loc].long)
                origin_1=(origin_lat_1,origin_long_1)
                if c_loc==b_loc or c_loc==a_loc:
                    continue
                if distance(origin_u,origin_1)<=70:
                    if c_loc in Dict3.get(a_loc):
                        conne=True
                if distance(origin_v,origin_1)<=70:
                    if c_loc in Dict3.get(b_loc):
                        conne=True
            if conne==True:
                continue   
                      
            R_count=R_count+1
            Routes['r'+str(R_count)]=[a_loc,b_loc]
                        
count2=ws.max_row
all_loc=[]
for i in df1.index:
    all_loc.append(df1['code'][i])
print(all_loc)
for k,v in Routes.items():
    sum=0
    count2=ws.max_row
    s=""
    list3=[]
    list3=v
    for u in range(0,len(list3)):
        s=s+list3[u]
        if u==len(list3)-1:
            continue
        s=s+"-"
    print(list3[0])
    print(list3[len(list3)-1])
    if list3[0] not in all_loc or list3[len(list3)-1] not in all_loc:
        continue   
    origin_lat_u=(float)(df1[df1['code']==list3[0]].lat)
    origin_long_u=(float)(df1[df1['code']==list3[0]].long)
    origin_u=(origin_lat_u,origin_long_u)
    origin_lat_v=(float)(df1[df1['code']==list3[len(list3)-1]].lat)
    origin_long_v=(float)(df1[df1['code']==list3[len(list3)-1]].long)
    origin_v=(origin_lat_v,origin_long_v)    
    count2=count2+1
    
    dist_list=[]
    dist_list=dist_map_api(origin_u,origin_v)
    ws['A'+str(count2)].value=str(k)
    ws['C'+str(count2)].value=list3[0]
    ws['D'+str(count2)].value=list3[len(list3)-1]
    
    ws['B'+str(count2)].value=s
    ws['G'+str(count2)].value=(int)(dist_list[1]/3600)
    ws['K'+str(count2)].value=(int)(dist_list[1]/3600)
    ws['I'+str(count2)].value=(int)(dist_list[0]/1000)
    ws['J'+str(count2)].value=(int)(dist_list[0]/1000)
wb.save("C:\\Users\\Nikhil\\Desktop\\dpWorld\\data\\Routes12.xlsx")
print("All_Routes= {")


for k, v in Routes.items():
    a=str(k)
    b=v
    print (a,':', b,',',sep='')
    

print('}')

all_loc_list1=[]
for i in df6.index:
    all_loc_list1.append(df6['Branches'][i])
print(all_loc_list1)
        
        